from flask import Blueprint, render_template, request, redirect, url_for, session, flash, current_app, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from .database import (create_session, get_session, get_sessions_by_user,
    save_target_allocations, get_target_allocations,
    add_stock, get_stocks, delete_stock,
    save_analysis_results, get_analysis_results,
    create_user, get_user_by_email, get_user_by_id)
from .analysis import run_analysis
import csv
import io

main = Blueprint("main", __name__)

def dbu():
    return current_app.config["DATABASE_URL"]

def dbt():
    return current_app.config["DB_TYPE"]

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            flash("Please log in to continue.", "warning")
            return redirect(url_for("main.login"))
        return f(*args, **kwargs)
    return decorated

@main.route("/")
def index():
    return render_template("index.html")

@main.route("/about")
def about():
    return render_template("about.html")

@main.route("/signup", methods=["GET", "POST"])
def signup():
    if session.get("user_id"):
        return redirect(url_for("main.index"))
    if request.method == "POST":
        name     = request.form.get("name", "").strip()
        email    = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        if not name or not email or not password:
            flash("All fields are required.", "danger")
            return render_template("signup.html")
        if len(password) < 6:
            flash("Password must be at least 6 characters.", "danger")
            return render_template("signup.html")
        existing = get_user_by_email(dbu(), dbt(), email)
        if existing:
            flash("An account with this email already exists. Please log in.", "danger")
            return render_template("signup.html")
        password_hash = generate_password_hash(password)
        user_id = create_user(dbu(), dbt(), name, email, password_hash)
        session["user_id"]  = user_id
        session["username"] = name
        flash(f"Welcome, {name}! Your account has been created.", "success")
        return redirect(url_for("main.index"))
    return render_template("signup.html")

@main.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user_id"):
        return redirect(url_for("main.index"))
    if request.method == "POST":
        email    = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        user = get_user_by_email(dbu(), dbt(), email)
        if not user or not check_password_hash(user["password_hash"], password):
            flash("Invalid email or password. Please try again.", "danger")
            return render_template("login.html")
        session["user_id"]  = user["id"]
        session["username"] = user["name"]
        flash(f"Welcome back, {user['name']}!", "success")
        return redirect(url_for("main.index"))
    return render_template("login.html")

@main.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "success")
    return redirect(url_for("main.login"))

@main.route("/new-session")
@login_required
def new_session():
    user_id  = session["user_id"]
    username = session.get("username", "User")
    sid = create_session(dbu(), dbt(), user_id=user_id, label=username)
    session["session_id"] = sid
    return redirect(url_for("main.target_allocation", session_id=sid))

@main.route("/session/<int:session_id>/targets", methods=["GET", "POST"])
@login_required
def target_allocation(session_id):
    existing = get_target_allocations(dbu(), dbt(), session_id)
    if request.method == "POST":
        sectors = request.form.getlist("sector")
        percentages = request.form.getlist("pct")
        allocations = []
        total = 0.0
        error = None
        for sector, pct in zip(sectors, percentages):
            sector = sector.strip()
            if not sector:
                continue
            try:
                pct_val = float(pct)
            except ValueError:
                error = f"'{pct}' is not a valid number."
                break
            if pct_val <= 0:
                error = "All percentages must be greater than 0."
                break
            total += pct_val
            allocations.append({"sector": sector, "target_pct": pct_val})
        if not error and len(allocations) == 0:
            error = "Please add at least one sector."
        if not error and round(total, 2) != 100.0:
            error = f"Your percentages add up to {total:.1f}%. They must total exactly 100%."
        if error:
            flash(error, "danger")
            return render_template("target_allocation.html", session_id=session_id, existing=existing)
        save_target_allocations(dbu(), dbt(), session_id, allocations)
        flash("Target allocations saved!", "success")
        return redirect(url_for("main.add_stocks", session_id=session_id))
    return render_template("target_allocation.html", session_id=session_id, existing=existing)

@main.route("/session/<int:session_id>/stocks", methods=["GET", "POST"])
@login_required
def add_stocks(session_id):
    targets = get_target_allocations(dbu(), dbt(), session_id)
    sectors = [t["sector"] for t in targets]
    if request.method == "POST":
        name      = request.form.get("name", "").strip()
        quantity  = request.form.get("quantity", "").strip()
        buy_price = request.form.get("buy_price", "").strip()
        sector    = request.form.get("sector", "").strip()
        error = None
        if not name:
            error = "Please enter the stock name."
        elif not quantity:
            error = "Please enter the quantity."
        elif not buy_price:
            error = "Please enter the buy price."
        elif not sector:
            error = "Please select a sector."
        else:
            try:
                qty_val   = float(quantity)
                price_val = float(buy_price)
                if qty_val <= 0:
                    error = "Quantity must be greater than 0."
                elif price_val <= 0:
                    error = "Buy price must be greater than 0."
            except ValueError:
                error = "Quantity and price must be numbers."
        if error:
            flash(error, "danger")
        else:
            add_stock(dbu(), dbt(), session_id,
                      name=name, quantity=float(quantity),
                      buy_price=float(buy_price), sector=sector)
            flash(f"'{name}' added successfully!", "success")
        return redirect(url_for("main.add_stocks", session_id=session_id))
    stocks = get_stocks(dbu(), dbt(), session_id)
    return render_template("add_stocks.html", session_id=session_id,
                           sectors=sectors, stocks=stocks)


def extract_rows_from_excel(file_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    NAME_ALIASES     = {"symbol", "stock name", "name", "scrip"}
    QUANTITY_ALIASES = {"quantity available", "quantity", "qty", "qty available", "available qty"}
    PRICE_ALIASES    = {"average price", "avg price", "buy price", "price", "avg. price", "average cost"}
    SECTOR_ALIASES   = {"sector"}
    header_row_index = None
    headers          = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_values = [str(v).strip().lower() if v is not None else "" for v in row]
        has_name   = any(v in NAME_ALIASES   for v in row_values)
        has_sector = any(v in SECTOR_ALIASES for v in row_values)
        if has_name and has_sector:
            header_row_index = row_idx
            headers = [str(v).strip() if v is not None else "" for v in row]
            break
    if not header_row_index:
        return []
    col_map = {}
    for col_idx, header in enumerate(headers):
        h = header.lower()
        if h in NAME_ALIASES:
            col_map[col_idx] = "Stock Name"
        elif h in QUANTITY_ALIASES:
            col_map[col_idx] = "Quantity Available"
        elif h in PRICE_ALIASES:
            col_map[col_idx] = "Average Price"
        elif h in SECTOR_ALIASES:
            col_map[col_idx] = "Sector"
    rows = []
    for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
        row_dict = {}
        for col_idx, value in enumerate(row):
            if col_idx in col_map:
                row_dict[col_map[col_idx]] = str(value).strip() if value is not None else ""
        if any(v for v in row_dict.values()):
            rows.append(row_dict)
    return rows


def extract_rows_from_csv(content):
    NAME_ALIASES     = {"symbol", "stock name", "name", "scrip"}
    QUANTITY_ALIASES = {"quantity available", "quantity", "qty", "qty available", "available qty"}
    PRICE_ALIASES    = {"average price", "avg price", "buy price", "price", "avg. price", "average cost"}
    SECTOR_ALIASES   = {"sector"}
    lines = content.splitlines()
    header_line_index = None
    for i, line in enumerate(lines):
        lower = line.lower()
        has_name   = any(alias in lower for alias in NAME_ALIASES)
        has_sector = any(alias in lower for alias in SECTOR_ALIASES)
        if has_name and has_sector:
            header_line_index = i
            break
    if header_line_index is None:
        reader = csv.DictReader(io.StringIO(content))
        return list(reader)
    data_content = "\n".join(lines[header_line_index:])
    reader = csv.DictReader(io.StringIO(data_content))
    rows = []
    for row in reader:
        normalised = {}
        for key, value in row.items():
            k = key.strip().lower()
            if k in NAME_ALIASES:
                normalised["Stock Name"] = str(value).strip()
            elif k in QUANTITY_ALIASES:
                normalised["Quantity Available"] = str(value).strip()
            elif k in PRICE_ALIASES:
                normalised["Average Price"] = str(value).strip()
            elif k in SECTOR_ALIASES:
                normalised["Sector"] = str(value).strip()
        if any(v for v in normalised.values()):
            rows.append(normalised)
    return rows


def parse_rows(rows, valid_sectors):
    added_rows = []
    errors     = []
    for i, row in enumerate(rows, start=2):
        name = (str(row.get("Stock Name", "") or row.get("stock_name", "") or row.get("Symbol", "") or row.get("Name", "")).strip())
        quantity = (str(row.get("Quantity Available", "") or row.get("Quantity", "") or row.get("quantity", "") or row.get("Qty", "")).strip())
        buy_price = (str(row.get("Average Price", "") or row.get("Buy Price", "") or row.get("buy_price", "") or row.get("Price", "")).strip())
        sector = (str(row.get("Sector", "") or row.get("sector", "")).strip())
        if not name and not quantity and not buy_price and not sector:
            continue
        if not name:
            errors.append(f"Row {i}: Missing stock name.")
            continue
        if not quantity:
            errors.append(f"Row {i}: Missing quantity for {name}.")
            continue
        if not buy_price:
            errors.append(f"Row {i}: Missing buy price for {name}.")
            continue
        if not sector:
            errors.append(f"Row {i}: Missing sector for {name}.")
            continue
        sector_match = next((s for s in valid_sectors if s.strip().lower() == sector.strip().lower()), None)
        if not sector_match:
            errors.append(f"Row {i} ({name}): Sector '{sector}' not found. Valid: {', '.join(valid_sectors)}")
            continue
        try:
            qty_val   = float(quantity)
            price_val = float(buy_price)
            if qty_val <= 0 or price_val <= 0:
                errors.append(f"Row {i}: Quantity and price must be greater than 0 for {name}.")
                continue
        except ValueError:
            errors.append(f"Row {i}: Quantity and price must be numbers for {name}.")
            continue
        added_rows.append({"name": name, "quantity": qty_val, "buy_price": price_val, "sector": sector_match})
    return added_rows, errors


@main.route("/session/<int:session_id>/stocks/upload", methods=["POST"])
@login_required
def upload_stocks(session_id):
    targets       = get_target_allocations(dbu(), dbt(), session_id)
    valid_sectors = [t["sector"] for t in targets]
    if "csv_file" not in request.files:
        flash("No file selected.", "danger")
        return redirect(url_for("main.add_stocks", session_id=session_id))
    file = request.files["csv_file"]
    if file.filename == "":
        flash("No file selected.", "danger")
        return redirect(url_for("main.add_stocks", session_id=session_id))
    filename = file.filename.lower()
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            rows = extract_rows_from_excel(file.read())
        elif filename.endswith(".csv"):
            rows = extract_rows_from_csv(file.read().decode("utf-8-sig"))
        else:
            flash("Please upload a CSV or Excel (.xlsx) file.", "danger")
            return redirect(url_for("main.add_stocks", session_id=session_id))
        if not rows:
            flash("Could not find any stock data in the file.", "danger")
            return redirect(url_for("main.add_stocks", session_id=session_id))
        valid_rows, errors = parse_rows(rows, valid_sectors)
        added = 0
        for r in valid_rows:
            add_stock(dbu(), dbt(), session_id, name=r["name"], quantity=r["quantity"], buy_price=r["buy_price"], sector=r["sector"])
            added += 1
        if added > 0:
            flash(f"Successfully imported {added} stock(s)!", "success")
        if errors:
            for err in errors[:5]:
                flash(err, "warning")
            if len(errors) > 5:
                flash(f"...and {len(errors) - 5} more errors.", "warning")
        if added == 0 and not errors:
            flash("No valid stocks found in the file.", "danger")
    except Exception as e:
        flash(f"Error reading file: {str(e)}", "danger")
    return redirect(url_for("main.add_stocks", session_id=session_id))


@main.route("/download-template")
@login_required
def download_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Stock Name", "Quantity", "Buy Price", "Sector"])
    writer.writerow(["Infosys", "50", "1450.00", "IT"])
    writer.writerow(["HDFC Bank", "100", "1620.00", "Banking"])
    writer.writerow(["Sun Pharma", "75", "1100.00", "Pharma"])
    writer.writerow(["Reliance", "30", "2800.00", "Energy"])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), mimetype="text/csv", as_attachment=True, download_name="portfolio_template.csv")


@main.route("/session/<int:session_id>/stocks/<int:stock_id>/delete", methods=["POST"])
@login_required
def remove_stock(session_id, stock_id):
    delete_stock(dbu(), dbt(), stock_id)
    flash("Stock removed.", "warning")
    return redirect(url_for("main.add_stocks", session_id=session_id))


@main.route("/session/<int:session_id>/analyze")
@login_required
def analyze(session_id):
    stocks  = get_stocks(dbu(), dbt(), session_id)
    targets = get_target_allocations(dbu(), dbt(), session_id)
    if not stocks:
        flash("Please add at least one stock before analyzing.", "danger")
        return redirect(url_for("main.add_stocks", session_id=session_id))
    results = run_analysis(stocks, targets)
    save_analysis_results(dbu(), dbt(), session_id, results)
    return redirect(url_for("main.results", session_id=session_id))


@main.route("/session/<int:session_id>/results")
@login_required
def results(session_id):
    analysis = get_analysis_results(dbu(), dbt(), session_id)
    stocks   = get_stocks(dbu(), dbt(), session_id)
    sess     = get_session(dbu(), dbt(), session_id)
    if not analysis:
        flash("No results found. Please run the analysis first.", "danger")
        return redirect(url_for("main.add_stocks", session_id=session_id))
    total_value = sum(float(s["quantity"]) * float(s["buy_price"]) for s in stocks if s["buy_price"])
    for row in analysis:
        row["actual_value"] = round((row["actual_pct"] / 100) * total_value, 2)
        row["target_value"] = round((row["target_pct"] / 100) * total_value, 2)
        row["gap_value"]    = round(row["actual_value"] - row["target_value"], 2)
        if row["status"] == "overweight":
            row["action"]        = "Reduce"
            row["action_amount"] = abs(row["gap_value"])
        elif row["status"] == "underweight":
            row["action"]        = "Increase"
            row["action_amount"] = abs(row["gap_value"])
        else:
            row["action"]        = "Hold"
            row["action_amount"] = 0
    return render_template("results.html", session_id=session_id, analysis=analysis, stocks=stocks, sess=sess, total_value=total_value)


@main.route("/history")
@login_required
def history():
    sessions = get_sessions_by_user(dbu(), dbt(), session["user_id"])
    return render_template("history.html", sessions=sessions)
